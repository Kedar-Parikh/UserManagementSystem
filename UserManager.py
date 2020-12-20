# This software is the property of Kedar Parikh
# Microsoft Excel is the Property of Microsoft Corporation
# Visit decryptingtechnology.com/contact for permissions
# For more information contact using email kedar@decryptingtechnology.com
# The usage of this software is free
# For distribution rights contact using email on kedar@decryptingtechnology.com

class LoginSystem:
    import openpyxl
    import random
    from time import sleep
    import base64
    import clipboard

    path = "Data.xlsx"
    workbook_object = openpyxl.load_workbook(path)
    sheet_object = workbook_object.active
    total_rows = 1048577

    def get_new_secret_num(self):
        while True:
            generated_num = self.random.randrange(2, self.total_rows)
            cell_object = self.sheet_object.cell(generated_num, 1)
            if not cell_object.value:
                return generated_num

    def start(self):
        try:
            while True:
                print('''
                         1) Login
                         2) SignUp
                         3) Delete Account
                         0) Exit
                                                        ''')
                menu_control = int(input('>> '))
                if menu_control == 1:
                    self.login()
                elif menu_control == 2:
                    self.signup()
                elif menu_control == 3:
                    self.delete_account()
                elif menu_control == 0:
                    exit(0)
                else:
                    print('[*] Invalid Input! Try Again')
        except ValueError:
            print('[*] Invalid input! Aborting....')
            exit(0)

    def login(self):
        secret_number = input('Secret Number: ')
        username = input('Username: ')
        password = input('Password: ')
        cell_object_username = self.sheet_object.cell(int(secret_number), 1)
        cellObjectPassword = self.sheet_object.cell(int(secret_number), 2)
        cellObjectData = self.sheet_object.cell(int(secret_number), 3)
        encrypted_password = self.encrypt(password)
        correct_password_value = cellObjectPassword.value
        if username == cell_object_username.value and str(encrypted_password, 'utf-8') == str(correct_password_value):
            print('[+] Login Successful!')
            data = self.base64.b64decode(cellObjectData.value, altchars=None)
            print('[+] Data Stored: \n' + str(data, 'utf-8'))
            self.sleep(2)
            print('[+] Logged out')
            self.start()
        else:
            print('[*] Invalid Username or Password')

    def signup(self):
        new_secret_num = self.get_new_secret_num()
        username = input('Username: ')
        password = input('Password: ')
        re_password = input('Re-Enter Password: ')
        data = ""
        if password == re_password and len(password) > 7:
            encrypted_password = self.encrypt(password)
            cell_name = self.sheet_object.cell(int(new_secret_num), 1)
            cell_password = self.sheet_object.cell(int(new_secret_num), 2)
            cell_data = self.sheet_object.cell(int(new_secret_num), 3)
            cell_name.value = username
            cell_password.value = str(encrypted_password, 'utf-8')
            self.workbook_object.save(self.path)
            print(f'Account with Username {username} created successfully!')
            print('Enter Data to be stored: ')
            data = input('>> ')
            encrypted_data = self.base64.b64encode(bytes(data, 'utf-8'), altchars=None)
            cell_data.value = encrypted_data
            self.workbook_object.save(self.path)
            print(f'[+] {new_secret_num}      This is your SECRET NUMBER DO NOT LOOSE IT')
            self.clipboard.copy(str(new_secret_num))
            print('[+] SECRET NUMBER COPIED TO CLIPBOARD')
        else:
            if password != re_password:
                print('[*] Passwords do not match!')
            else:
                print('[*] Password length is too short! Try Again')

    def delete_account(self):
        secret_number = input('Enter Secret Number: ')
        username = input('Enter Username: ')
        password = input('Enter Password: ')
        cell_object_username = self.sheet_object.cell(int(secret_number), 1)
        cell_object_password = self.sheet_object.cell(int(secret_number), 2)
        cell_object_data = self.sheet_object.cell(int(secret_number), 3)
        encrypted_password = self.encrypt(password)
        actual_password_value = cell_object_password.value
        if username == cell_object_username.value and str(encrypted_password, 'utf-8') == actual_password_value:
            print('[*] Deleting Account Information and Data')
            cell_object_username.value = ""
            cell_object_password.value = ""
            cell_object_data.value = ""
            self.workbook_object.save(self.path)
            print('[*] Account Information and Data Deleted Successfully')
        else:
            print('[*] Incorrect Username or Password!')

    def encrypt(self, data):
        password_length = len(data)
        hashed_password = ""

        if password_length % 2 == 0:
            password_length_type = 'even'
        else:
            password_length_type = 'odd'

        if password_length_type == 'even':
            hashes_required = (password_length / 2) - 1

        else:
            hashes_required = (password_length - 1) / 2

        characters_transferred = 0
        hashes_completed = 0
        while hashes_completed != hashes_required:
            for character in data:
                characters_transferred = characters_transferred + 1
                if characters_transferred % 2 == 0:
                    hashed_password = hashed_password + str(character) + '~`^!'
                    hashes_completed = hashes_completed + 1
                elif characters_transferred % 2 != 0:
                    hashed_password = hashed_password + str(character)

        encrypted_password = self.base64.b64encode(bytes(hashed_password, 'utf-8'), altchars=None)
        return encrypted_password
